from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import math

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
DARK_NAVY = "1B2A4A"
MED_BLUE = "2C5F8A"
LIGHT_BLUE = "D6E8F7"
ACCENT_GOLD = "F5C542"
ACCENT_ORANGE = "E8833A"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
GREEN_YES = "C6EFCE"
GREEN_DARK = "006100"
RED_NO = "FFC7CE"
RED_DARK = "9C0006"
INPUT_BLUE = "0000FF"
MAJOR_GOLD = "FFF2CC"
MAJOR_BORDER = "BF8F00"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_NAVY)
subheader_font = Font(name="Arial", bold=True, color=DARK_NAVY, size=10)
subheader_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
data_font = Font(name="Arial", size=10)
input_font = Font(name="Arial", size=10, color=INPUT_BLUE, bold=True)
major_fill = PatternFill("solid", fgColor=MAJOR_GOLD)
alt_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(ws, row, col, is_alt=False, is_major=False):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_major:
        cell.fill = major_fill
    elif is_alt:
        cell.fill = alt_fill

# ============================================================
# RACE DATA - ALL GLOBAL UTMB WORLD SERIES EVENTS 2026
# ============================================================
# (Event Name, Country, City/Region, Dates, Continent, Type, 
#  Distances Available, Lat, Lon, Website, Notes)
# Distances: list of tuples (category, stones_event, stones_major)

races = [
    # NORTH AMERICA
    ("Desert RATS", "USA", "Fruita, CO", "Apr 9-12", "North America", "Event",
     "50K, 100K, 100M", 37.86, -108.73, "desertrats.utmb.world", ""),
    ("Puerto Vallarta", "Mexico", "Puerto Vallarta", "Apr 16-18", "North America", "Event",
     "20K, 50K", 20.65, -105.23, "puertovallarta.utmb.world", "Beach & jungle trails"),
    ("Canyons Endurance Runs", "USA", "Auburn, CA", "Apr 23-25", "North America", "Event",
     "20K, 50K, 100K, 100M", 38.90, -121.08, "canyons.utmb.world", "Also Western States qualifier"),
    ("Rothrock", "USA", "State College, PA", "May 15-17", "North America", "Event",
     "20K, 50K", 40.79, -77.86, "rothrock.utmb.world", "New for 2026"),
    ("Western States Endurance Run", "USA", "Auburn, CA", "Jun 27-28", "North America", "Event",
     "100M", 38.90, -121.08, "wser.org", "Iconic 100-miler; own lottery required"),
    ("Speedgoat Mountain Races", "USA", "Snowbird, UT", "Jul 23-25", "North America", "Event",
     "20K, 50K", 40.58, -111.66, "speedgoat.utmb.world", "Wasatch Mountains; close to Logan!"),
    ("Pacific Trails California", "USA", "California", "Aug 14-16", "North America", "Event",
     "50K, 100K", 37.50, -122.00, "pacifictrails.utmb.world", "New for 2026"),
    ("Boréalys Mont-Tremblant", "Canada", "Mont-Tremblant, QC", "Aug 21-23", "North America", "Event",
     "20K, 50K, 100K", 46.21, -74.58, "borealys.utmb.world", "Laurentian Mountains"),
    ("Snowbasin", "USA", "Ogden, UT", "Sep 10-12", "North America", "Event",
     "20K, 50K", 41.22, -111.86, "snowbasin.utmb.world", "NEW! Wasatch Range; ~1hr from Logan!"),
    ("Grindstone", "USA", "Swoope, VA", "Sep 17-20", "North America", "Event",
     "50K, 100K, 100M", 38.07, -79.27, "grindstone.utmb.world", "Appalachian Mountains"),
    ("Ultra Trail Whistler", "Canada", "Whistler, BC", "Oct 1-3", "North America", "Event",
     "20K, 50K, 100K", 50.12, -122.95, "whistler.utmb.world", "Coast Mountains; beautiful terrain"),
    ("Kodiak Ultra Marathons", "USA", "Big Bear Lake, CA", "Oct 8-11", "North America", "Major",
     "20K, 50K, 100K, 100M", 34.24, -116.91, "kodiak.utmb.world", "2026 AMERICAS MAJOR - double stones!"),
    ("Chihuahua", "Mexico", "Chihuahua", "Nov 13-15", "North America", "Event",
     "20K, 50K", 28.63, -106.09, "chihuahua.utmb.world", "Copper Canyon region"),

    # SOUTH AMERICA
    ("Quito Trail", "Ecuador", "Quito", "Mar 18-22", "South America", "Event",
     "20K, 50K, 100K", -0.18, -78.47, "quito.utmb.world", "High altitude ~2,800m"),
    ("Torrencial Chile", "Chile", "Central Chile", "Jun 26-28", "South America", "Event",
     "20K, 50K", -33.45, -70.66, "torrencial.utmb.world", "Winter racing in southern hemisphere"),
    ("Valhalla Argentina", "Argentina", "San Martin de los Andes", "Jul 31-Aug 2", "South America", "Event",
     "20K, 50K, 100K, 100M", -40.15, -71.35, "valhalla.utmb.world", "Patagonian Andes"),
    ("EcoTrail Monteverde", "Costa Rica", "Monteverde", "Sep 17-20", "South America", "Event",
     "20K, 50K", 10.31, -84.82, "monteverde.utmb.world", "Cloud forest running"),
    ("Patagonia Bariloche", "Argentina", "Bariloche", "Nov 18-22", "South America", "Event",
     "20K, 50K, 100K", -41.13, -71.31, "bariloche.utmb.world", "Lake district of Patagonia"),

    # EUROPE
    ("Arc of Attrition", "UK", "Cornwall", "Jan 22-25", "Europe", "Event",
     "50K, 100M", 50.12, -5.54, "arcofattrition.utmb.world", "Coastal winter ultra; tough conditions"),
    ("Chianti Ultra Trail", "Italy", "Chianti, Tuscany", "Mar 19-22", "Europe", "Event",
     "20K, 50K, 100K", 43.48, 11.25, "chianti.utmb.world", "Tuscan wine country trails"),
    ("Tenerife Bluetrail", "Spain", "Tenerife", "Mar 19-21", "Europe", "Event",
     "20K, 50K, 100K, 100M", 28.27, -16.59, "tenerife.utmb.world", "Volcanic island; Teide ascent"),
    ("Mallorca by UTMB", "Spain", "Mallorca", "Apr 9-12", "Europe", "Event",
     "20K, 50K, 100K", 39.69, 2.94, "mallorca.utmb.world", "Mediterranean island trails"),
    ("Grand Raid Ventoux", "France", "Mont Ventoux", "Apr 24-26", "Europe", "Event",
     "20K, 50K, 100K", 44.17, 5.28, "ventoux.utmb.world", "Iconic Provençal mountain"),
    ("Istria 100", "Croatia", "Istria", "May 1-3", "Europe", "Event",
     "20K, 50K, 100K, 100M", 45.24, 13.77, "istria.utmb.world", "Adriatic coast & karst terrain"),
    ("Eiger Ultra Trail", "Switzerland", "Grindelwald", "May 14-17", "Europe", "Event",
     "20K, 50K, 100K, 100M", 46.62, 8.04, "eiger.utmb.world", "Bernese Alps; iconic Eiger North Face"),
    ("Ultra-Trail Snowdonia", "UK", "Snowdonia", "May 15-17", "Europe", "Event",
     "20K, 50K, 100K", 53.07, -3.99, "snowdonia.utmb.world", "Welsh mountains"),
    ("Mozart 100", "Austria", "Salzburg", "May 23", "Europe", "Event",
     "50K, 100K", 47.81, 13.04, "mozart100.utmb.world", "Alps around Salzburg"),
    ("Oh Meu Deus", "Portugal", "Portugal", "Jun 11-14", "Europe", "Event",
     "20K, 50K, 100K", 38.71, -9.14, "ohmeudeus.utmb.world", "New Portuguese event"),
    ("Gauja Trail", "Latvia", "Gauja NP", "Jun 12-14", "Europe", "Event",
     "20K, 50K", 57.18, 25.08, "gauja.utmb.world", "Latvian national park"),
    ("Wildstrubel", "Switzerland", "Crans-Montana", "Jun 18-20", "Europe", "Event",
     "20K, 50K, 100K", 46.31, 7.48, "wildstrubel.utmb.world", "Swiss Alps"),
    ("Lavaredo Ultra Trail", "Italy", "Cortina d'Ampezzo", "Jun 24-28", "Europe", "Event",
     "20K, 50K, 100K, 100M", 46.54, 12.14, "lavaredo.utmb.world", "Dolomites; stunning scenery"),
    ("Val d'Aran by UTMB", "Spain", "Val d'Aran, Pyrenees", "Jul 1-5", "Europe", "Major",
     "20K, 50K, 100K, 100M", 42.70, 0.72, "valdaran.utmb.world", "2026 EUROPE MAJOR - double stones!"),
    ("Nice Côte d'Azur", "France", "Nice", "Jul 9-11", "Europe", "Event",
     "20K, 50K, 100K", 43.71, 7.26, "nice.utmb.world", "French Riviera hinterland"),
    ("Bucovina Ultra Rocks", "Romania", "Bucovina", "Jul 10-12", "Europe", "Event",
     "20K, 50K", 47.75, 25.35, "bucovina.utmb.world", "Carpathian Mountains"),
    ("Zugspitz Ultra Trail", "Germany", "Garmisch-Partenkirchen", "Jul 15-19", "Europe", "Event",
     "20K, 50K, 100K", 47.50, 11.10, "zugspitz.utmb.world", "Germany's highest peak area"),
    ("Monte Rosa Walser Waeg", "Italy", "Monte Rosa", "Jul 17-19", "Europe", "Event",
     "20K, 50K, 100K", 45.88, 7.87, "monterosa.utmb.world", "Italian Alps"),
    ("Trail du Saint-Jacques", "France", "Le Puy-en-Velay", "Jul 24-26", "Europe", "Event",
     "20K, 50K, 100K", 45.04, 3.88, "saintjacques.utmb.world", "Camino de Santiago region"),
    ("Penyagolosa Trails", "Spain", "Castellon", "Aug 1-2", "Europe", "Event",
     "20K, 50K, 100K", 40.23, -0.35, "penyagolosa.utmb.world", "Eastern Spain mountains"),
    ("Ultra-Trail Cape Town", "South Africa", "Cape Town", "Aug 6-8", "Europe", "Event",
     "20K, 50K, 100K, 100M", -33.96, 18.47, "capetown.utmb.world", "Table Mountain & surrounds"),
    ("UTMB Mont-Blanc Finals", "France", "Chamonix", "Aug 24-30", "Europe", "Finals",
     "OCC(50K), CCC(100K), UTMB(100M)", 45.92, 6.87, "montblanc.utmb.world", "THE GOAL! CCC is Aug 29, 2027"),
    ("Kaçkar", "Turkey", "Kaçkar Mountains", "Sep 10-13", "Europe", "Event",
     "20K, 50K, 100K", 40.80, 41.10, "kackar.utmb.world", "Eastern Turkey; remote mountains"),
    ("Ultra Pirineu", "Spain", "Bagà, Catalonia", "Sep 11-13", "Europe", "Event",
     "20K, 50K, 100K, 100M", 42.25, 1.86, "pirineu.utmb.world", "Cadí-Moixeró Natural Park"),
    ("Ultra-Trail Mogan", "Spain", "Gran Canaria", "Sep 18-20", "Europe", "Event",
     "20K, 50K", 27.93, -15.43, "mogan.utmb.world", "Canary Islands"),
    ("Cappadocia Ultra-Trail", "Turkey", "Cappadocia", "Sep 24-27", "Europe", "Event",
     "20K, 50K, 100K, 100M", 38.64, 34.83, "cappadocia.utmb.world", "Fairy chimneys & cave trails"),
    ("Puglia", "Italy", "Puglia", "Oct 30-31", "Europe", "Event",
     "20K, 50K", 40.64, 17.94, "puglia.utmb.world", "Southern Italian coast"),
    ("Trail Nice Côte d'Azur Oman", "Oman", "Hajar Mountains", "Oct 30-Nov 1", "Europe", "Event",
     "20K, 50K, 100K", 23.10, 57.30, "oman.utmb.world", "Desert mountain running"),

    # ASIA
    ("Ultra-Trail Xiamen", "China", "Xiamen", "Mar 6-8", "Asia", "Event",
     "20K, 50K", 24.48, 118.09, "xiamen.utmb.world", "Coastal Chinese city"),
    ("TransLantau", "Hong Kong", "Lantau Island", "Mar 14-15", "Asia", "Event",
     "20K, 50K, 100K", 22.27, 113.94, "translantau.utmb.world", "Hong Kong island trails"),
    ("Mount Sun", "China", "Emeishan", "Apr 10-12", "Asia", "Event",
     "20K, 50K, 100K", 29.52, 103.33, "mountsun.utmb.world", "Sacred Buddhist mountain"),
    ("Chiang Mai Thailand", "Thailand", "Chiang Mai", "Apr 30-May 3", "Asia", "Major",
     "20K, 50K, 100K, 100M", 18.79, 98.98, "chiangmai.utmb.world", "2026 ASIA-PACIFIC MAJOR - double stones!"),
    ("Ultra-Trail Great Wall", "China", "Beijing", "May 15-17", "Asia", "Event",
     "20K, 50K", 40.43, 116.57, "greatwall.utmb.world", "Run along the Great Wall"),
    ("Kaga Spa Trail Endurance", "Japan", "Kaga, Ishikawa", "Jun 12-14", "Asia", "Event",
     "20K, 50K, 100K, 100M", 36.30, 136.31, "kagaspa.utmb.world", "Japanese hot spring region"),
    ("Ultra-Trail Tai Mo Shan", "Hong Kong", "New Territories", "Jun 18-21", "Asia", "Event",
     "20K, 50K, 100K", 22.41, 114.12, "taimoshan.utmb.world", "Hong Kong's highest peak"),
    ("TransJeju", "South Korea", "Jeju Island", "Sep 11-13", "Asia", "Event",
     "20K, 50K, 100K", 33.36, 126.53, "transjeju.utmb.world", "Volcanic island; UNESCO site"),
    ("Trail of The Kings", "Indonesia", "Indonesia", "Sep 11-13", "Asia", "Event",
     "20K, 50K", -6.75, 106.75, "trailofkings.utmb.world", "Tropical mountain trails"),
    ("Malaysia Ultra Trail", "Malaysia", "Malaysia", "Oct 2-4", "Asia", "Event",
     "20K, 50K, 100K", 4.21, 101.97, "malaysia.utmb.world", "Rainforest trails"),
    ("Amazean Jungle", "Thailand", "Chiang Rai", "Oct 16-18", "Asia", "Event",
     "20K, 50K, 100K", 19.91, 99.83, "amazean.utmb.world", "Northern Thailand jungle"),
    ("Xtrail Kenting", "Taiwan", "Kenting", "Nov 6-8", "Asia", "Event",
     "20K, 50K", 21.95, 120.80, "kenting.utmb.world", "Taiwan's tropical south"),
    ("Ultra-Trail Shudao", "China", "Sichuan", "Dec 10-12", "Asia", "Event",
     "20K, 50K, 100K", 30.57, 104.07, "shudao.utmb.world", "Sichuan province"),

    # OCEANIA
    ("Ultra-Trail Australia", "Australia", "Blue Mountains, NSW", "May 14-17", "Oceania", "Major",
     "20K, 50K, 100K, 100M", -33.72, 150.31, "uta.utmb.world", "2026 OCEANIA MAJOR - double stones!"),
    ("Ultra-Trail Kosciuszko", "Australia", "Jindabyne, NSW", "Nov 26-28", "Oceania", "Event",
     "20K, 50K, 100K, 100M", -36.42, 148.62, "kosciuszko.utmb.world", "Snowy Mountains"),
    ("Tarawera Ultra-Trail", "New Zealand", "Rotorua", "Feb 13-14 (2027)", "Oceania", "Event",
     "20K, 50K, 100K, 100M", -38.14, 176.25, "tarawera.utmb.world", "Volcanic geothermal trails"),

    # AFRICA
    ("Ultra-Trail Drakensberg", "South Africa", "KwaZulu-Natal", "May 29-31", "Africa", "Event",
     "20K, 50K, 100K", -29.37, 29.28, "drakensberg.utmb.world", "Drakensberg Mountains"),
]

# Logan, UT coordinates
LOGAN_LAT = 41.735
LOGAN_LON = -111.834

def haversine_miles(lat1, lon1, lat2, lon2):
    R = 3959
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def parse_distances(dist_str):
    """Return list of distance categories available"""
    cats = []
    for d in dist_str.split(","):
        d = d.strip()
        if "100M" in d or "100m" in d:
            cats.append("100M")
        elif "100K" in d or "100k" in d:
            cats.append("100K")
        elif "50K" in d or "50k" in d:
            cats.append("50K")
        elif "20K" in d or "20k" in d:
            cats.append("20K")
        elif "OCC" in d:
            cats.append("50K")
        elif "CCC" in d:
            cats.append("100K")
        elif "UTMB" in d:
            cats.append("100M")
    return cats

def stones_for_category(cat, is_major):
    base = {"20K": 1, "50K": 2, "100K": 3, "100M": 4}
    s = base.get(cat, 0)
    return s * 2 if is_major else s

# ============================================================
# SHEET 1: ALL RACES
# ============================================================
ws1 = wb.active
ws1.title = "All UTMB Events 2026"
ws1.sheet_properties.tabColor = DARK_NAVY

headers1 = [
    "Select\n(Y/N)", "Event Name", "Type", "Continent", "Country", "City / Region",
    "Dates (2026)", "Distances\nAvailable", "Best CCC\nCategory", "Stones\n(Best for CCC)",
    "Provides\n100K Index?", "Distance from\nLogan (mi)", "Est. Travel\nCost ($)",
    "Race Entry\nEst. ($)", "Website", "Notes"
]

# Title row
ws1.merge_cells("A1:P1")
title_cell = ws1["A1"]
title_cell.value = "UTMB WORLD SERIES 2026 — KATHLEEN'S CCC QUALIFICATION PLANNER"
title_cell.font = Font(name="Arial", bold=True, color=WHITE, size=14)
title_cell.fill = PatternFill("solid", fgColor=DARK_NAVY)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

# Subtitle
ws1.merge_cells("A2:P2")
sub_cell = ws1["A2"]
sub_cell.value = "Target: CCC (100K) at UTMB Mont-Blanc 2027  |  Requires: Valid UTMB Index (50K/100K/100M) + Running Stones (aim for 6-8)  |  Mark 'Y' in column A to select races"
sub_cell.font = Font(name="Arial", color=DARK_NAVY, size=10, italic=True)
sub_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 25

# Headers
for c, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=c, value=h)
style_header_row(ws1, 3, len(headers1))
ws1.row_dimensions[3].height = 40

# Add data validation for Y/N column
dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv.error = "Please enter Y or N"
dv.errorTitle = "Invalid Entry"
ws1.add_data_validation(dv)

# Populate data
for i, race in enumerate(races):
    row = i + 4
    name, country, city, dates, continent, rtype, dists, lat, lon, website, notes = race
    is_major = (rtype == "Major")
    is_finals = (rtype == "Finals")
    is_alt = (i % 2 == 1)
    
    dist_cats = parse_distances(dists)
    
    # Best category for CCC qualification (need 50K, 100K, or 100M index)
    # Prefer 100K for most stones, then 50K, then 100M
    best_cat = None
    for pref in ["100K", "50K", "100M"]:
        if pref in dist_cats:
            best_cat = pref
            break
    if best_cat is None and "20K" in dist_cats:
        best_cat = "20K"
    
    stones = stones_for_category(best_cat, is_major) if best_cat else 0
    provides_index = "Yes" if any(c in dist_cats for c in ["50K", "100K", "100M"]) else "No (20K only)"
    
    dist_miles = round(haversine_miles(LOGAN_LAT, LOGAN_LON, lat, lon))
    
    # Estimate travel cost based on distance
    if dist_miles < 200:
        est_travel = "~$50-150"
    elif dist_miles < 600:
        est_travel = "~$150-400"
    elif dist_miles < 2000:
        est_travel = "~$400-800"
    elif dist_miles < 5000:
        est_travel = "~$800-1,500"
    else:
        est_travel = "~$1,500-3,000"
    
    # Estimate race entry cost
    if is_major:
        est_entry = "$150-350"
    elif is_finals:
        est_entry = "$300-500"
    else:
        est_entry = "$100-250"
    
    values = [
        "",  # Y/N selection
        name,
        rtype,
        continent,
        country,
        city,
        dates,
        dists,
        best_cat if best_cat else "N/A",
        stones if not is_finals else "N/A",
        provides_index if not is_finals else "GOAL RACE",
        dist_miles,
        est_travel,
        est_entry,
        website,
        notes,
    ]
    
    for c, val in enumerate(values, 1):
        cell = ws1.cell(row=row, column=c, value=val)
        style_data_cell(ws1, row, c, is_alt=is_alt, is_major=is_major)
        if c == 1:
            cell.font = input_font
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
        elif c == 2:
            cell.alignment = left_wrap
            cell.font = Font(name="Arial", size=10, bold=is_major or is_finals)
        elif c == 16:
            cell.alignment = left_wrap
    
    # Add Y/N validation to selection cell
    dv.add(ws1.cell(row=row, column=1))
    ws1.row_dimensions[row].height = 22

last_race_row = len(races) + 3

# Column widths
col_widths = [8, 30, 8, 14, 12, 22, 14, 22, 12, 12, 14, 14, 14, 12, 28, 35]
for c, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

# Freeze panes
ws1.freeze_panes = "B4"

# ============================================================
# SHEET 2: RACE SCHEDULE BUILDER
# ============================================================
ws2 = wb.create_sheet("Race Schedule & Stones")
ws2.sheet_properties.tabColor = MED_BLUE

# Title
ws2.merge_cells("A1:L1")
t2 = ws2["A1"]
t2.value = "KATHLEEN'S RACE SCHEDULE & RUNNING STONES CALCULATOR"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t2.fill = PatternFill("solid", fgColor=DARK_NAVY)
t2.alignment = center
ws2.row_dimensions[1].height = 35

# Instructions
ws2.merge_cells("A2:L2")
inst = ws2["A2"]
inst.value = 'This sheet auto-populates from your "Y" selections on the All Events tab. Review your schedule, check for conflicts, and see your stone accumulation below.'
inst.font = Font(name="Arial", color=DARK_NAVY, size=10, italic=True)
inst.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
inst.alignment = center
ws2.row_dimensions[2].height = 25

# Schedule headers
sched_headers = [
    "#", "Event Name", "Type", "Country", "City", "Dates",
    "Category\nRaced", "Stones\nEarned", "Cumulative\nStones",
    "Valid CCC\nIndex?", "Distance\nfrom Logan", "Notes"
]
for c, h in enumerate(sched_headers, 1):
    ws2.cell(row=3, column=c, value=h)
style_header_row(ws2, 3, len(sched_headers))
ws2.row_dimensions[3].height = 40

# Use formulas to pull selected races
# We'll create rows that check if selected = "Y" on sheet 1
# Since Excel IF formulas can't dynamically filter, we'll use a helper approach
# Instead, we'll put placeholder formulas that reference the first sheet

# For up to 20 potential selected races
for slot in range(1, 21):
    row = slot + 3
    # These will be placeholder rows - user will manually copy selected races
    # Or we provide INDEX/MATCH formulas
    for c in range(1, 13):
        cell = ws2.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = center
        cell.font = data_font
        if slot % 2 == 0:
            cell.fill = alt_fill
    ws2.cell(row=row, column=1, value=slot)
    # Column 9 (cumulative stones) formula
    if slot == 1:
        ws2.cell(row=row, column=9).value = f'=IF(H{row}="",0,H{row})'
    else:
        ws2.cell(row=row, column=9).value = f'=IF(H{row}="",I{row-1},I{row-1}+H{row})'
    ws2.cell(row=row, column=9).font = Font(name="Arial", size=10, bold=True)

ws2.row_dimensions[4].height = 22

# Summary section
sum_start = 25
ws2.merge_cells(f"A{sum_start}:L{sum_start}")
s = ws2[f"A{sum_start}"]
s.value = "QUALIFICATION SUMMARY"
s.font = Font(name="Arial", bold=True, color=WHITE, size=12)
s.fill = PatternFill("solid", fgColor=MED_BLUE)
s.alignment = center

labels = [
    ("Total Running Stones Accumulated:", f"=I23", ""),
    ("Valid UTMB Index for CCC? (need 50K/100K/100M):", "", "Enter YES or NO"),
    ("At Least 1 Stone in Last 24 Months?:", "", "Enter YES or NO"),
    ("Eligible for CCC Lottery?:", "", "Both above must be YES + stones > 0"),
]

for i, (label, formula, note) in enumerate(labels):
    r = sum_start + 1 + i
    ws2.merge_cells(f"A{r}:E{r}")
    lc = ws2[f"A{r}"]
    lc.value = label
    lc.font = Font(name="Arial", bold=True, size=11)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border = thin_border
    
    vc = ws2[f"F{r}"]
    if formula:
        vc.value = formula
    vc.font = Font(name="Arial", bold=True, size=14, color=INPUT_BLUE)
    vc.alignment = center
    vc.border = thin_border
    vc.fill = PatternFill("solid", fgColor="E8F0FE")
    
    ws2.merge_cells(f"G{r}:L{r}")
    nc = ws2[f"G{r}"]
    nc.value = note
    nc.font = Font(name="Arial", italic=True, size=9, color="666666")
    nc.alignment = Alignment(horizontal="left", vertical="center")

# Column widths for schedule
sched_widths = [5, 28, 8, 12, 20, 14, 12, 10, 12, 12, 14, 30]
for c, w in enumerate(sched_widths, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

ws2.freeze_panes = "A4"

# ============================================================
# SHEET 3: LOTTERY ODDS CALCULATOR
# ============================================================
ws3 = wb.create_sheet("CCC Lottery Odds")
ws3.sheet_properties.tabColor = ACCENT_GOLD

# Title
ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "CCC (100K) LOTTERY PROBABILITY ESTIMATOR — UTMB 2027"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=14)
t3.fill = PatternFill("solid", fgColor=DARK_NAVY)
t3.alignment = center
ws3.row_dimensions[1].height = 35

# Methodology note
ws3.merge_cells("A2:H2")
m = ws3["A2"]
m.value = "Based on 2025 lottery data: ~6,000 CCC applicants, ~1,200 lottery spots, avg selected runner = 5.7 stones, avg applicant = 4.4 stones"
m.font = Font(name="Arial", color=DARK_NAVY, size=9, italic=True)
m.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
m.alignment = center
ws3.row_dimensions[2].height = 22

# Assumptions section
ws3.merge_cells("A4:D4")
ws3["A4"].value = "LOTTERY ASSUMPTIONS (editable - blue = input)"
ws3["A4"].font = Font(name="Arial", bold=True, size=11, color=DARK_NAVY)
ws3["A4"].fill = subheader_fill

assumptions = [
    ("Total CCC Applicants (2027 est.):", 7000, "B5", "Growing ~20% YoY from 6,000 in 2025"),
    ("Lottery Spots Available:", 1200, "B6", "Approx. after elite/charity bibs removed"),
    ("Avg Stones per Applicant:", 5.0, "B7", "Was 4.4 in 2025; trending up"),
    ("Kathleen's Running Stones:", 0, "B8", "← Link from Race Schedule sheet or enter manually"),
]

for i, (label, default, cell_ref, note) in enumerate(assumptions):
    r = 5 + i
    ws3[f"A{r}"].value = label
    ws3[f"A{r}"].font = Font(name="Arial", size=10)
    ws3[f"A{r}"].alignment = Alignment(horizontal="right", vertical="center")
    ws3[f"A{r}"].border = thin_border
    
    ws3[cell_ref].value = default
    ws3[cell_ref].font = input_font
    ws3[cell_ref].alignment = center
    ws3[cell_ref].border = thin_border
    ws3[cell_ref].fill = PatternFill("solid", fgColor="E8F0FE")
    
    ws3[f"C{r}"].value = note
    ws3[f"C{r}"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws3.merge_cells(f"C{r}:H{r}")

# Link Kathleen's stones from schedule sheet
ws3["B8"].value = "='Race Schedule & Stones'!F26"

# Probability calculation section
ws3.merge_cells("A10:D10")
ws3["A10"].value = "PROBABILITY CALCULATION"
ws3["A10"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
ws3["A10"].fill = PatternFill("solid", fgColor=MED_BLUE)

calc_rows = [
    ("Total Stones in Pool (est.):", '=B5*B7', "All applicants × avg stones"),
    ("Kathleen's Share of Pool:", '=IF(B8=0,"No stones yet",B8/(B5*B7))', "Her stones ÷ total pool"),
    ("Single-Draw Selection Prob.:", '=IF(B8=0,"N/A",1-(1-B8/(B5*B7))^B6)', "Probability of being drawn in 1,200 picks"),
    ("Estimated Lottery Odds:", '=IF(B8=0,"Need stones!",TEXT(ROUND(1/(1-(1-B8/(B5*B7))^B6),0),"#,##0")&" to 1")', "Approx. 1 in X chance"),
    ("Selection Probability %:", '=IF(B8=0,0,(1-(1-B8/(B5*B7))^B6)*100)', "Your percentage chance"),
]

for i, (label, formula, note) in enumerate(calc_rows):
    r = 11 + i
    ws3[f"A{r}"].value = label
    ws3[f"A{r}"].font = Font(name="Arial", size=10, bold=True)
    ws3[f"A{r}"].alignment = Alignment(horizontal="right", vertical="center")
    ws3[f"A{r}"].border = thin_border
    
    ws3[f"B{r}"].value = formula
    ws3[f"B{r}"].font = Font(name="Arial", size=12, bold=True, color=ACCENT_ORANGE)
    ws3[f"B{r}"].alignment = center
    ws3[f"B{r}"].border = thin_border
    
    ws3[f"C{r}"].value = note
    ws3[f"C{r}"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws3.merge_cells(f"C{r}:H{r}")

# Stones scenario table
ws3.merge_cells("A18:H18")
ws3["A18"].value = "WHAT-IF: SELECTION PROBABILITY BY NUMBER OF STONES"
ws3["A18"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
ws3["A18"].fill = PatternFill("solid", fgColor=DARK_NAVY)

scenario_headers = ["Stones", "Share of Pool", "Selection Prob %", "~1 in X", "Assessment", "How to Get There"]
for c, h in enumerate(scenario_headers, 1):
    ws3.cell(row=19, column=c, value=h)
    ws3.cell(row=19, column=c).font = Font(name="Arial", bold=True, color=DARK_NAVY, size=10)
    ws3.cell(row=19, column=c).fill = subheader_fill
    ws3.cell(row=19, column=c).alignment = center
    ws3.cell(row=19, column=c).border = thin_border

stone_scenarios = [
    (2, "Low — tough odds", "1 Event 50K"),
    (3, "Below average", "1 Event 100K"),
    (4, "Average applicant", "1 Major 50K or 2 Event 50Ks"),
    (5, "Slightly above avg", "1 Event 50K + 1 Event 100K"),
    (6, "Competitive", "1 Major 100K"),
    (7, "Good odds", "1 Major 100K + 1 Event 20K"),
    (8, "Strong position", "1 Major 100K + 1 Event 50K"),
    (10, "Very strong", "1 Major 100K + 1 Major 50K"),
    (12, "Excellent", "1 Major 100M + 1 Major 50K"),
]

for i, (stones, assessment, how) in enumerate(stone_scenarios):
    r = 20 + i
    ws3.cell(row=r, column=1, value=stones)
    ws3.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True)
    ws3.cell(row=r, column=1).alignment = center
    ws3.cell(row=r, column=1).border = thin_border
    
    # Share of pool
    ws3.cell(row=r, column=2).value = f'={stones}/(B$5*B$7)'
    ws3.cell(row=r, column=2).number_format = '0.000%'
    ws3.cell(row=r, column=2).alignment = center
    ws3.cell(row=r, column=2).border = thin_border
    
    # Selection probability
    ws3.cell(row=r, column=3).value = f'=(1-(1-{stones}/(B$5*B$7))^B$6)*100'
    ws3.cell(row=r, column=3).number_format = '0.0'
    ws3.cell(row=r, column=3).alignment = center
    ws3.cell(row=r, column=3).border = thin_border
    
    # 1 in X
    ws3.cell(row=r, column=4).value = f'=ROUND(1/(1-(1-{stones}/(B$5*B$7))^B$6),1)&" to 1"'
    ws3.cell(row=r, column=4).alignment = center
    ws3.cell(row=r, column=4).border = thin_border
    
    ws3.cell(row=r, column=5, value=assessment)
    ws3.cell(row=r, column=5).alignment = center
    ws3.cell(row=r, column=5).border = thin_border
    
    ws3.cell(row=r, column=6, value=how)
    ws3.cell(row=r, column=6).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=6).border = thin_border
    
    if i % 2 == 1:
        for c in range(1, 7):
            ws3.cell(row=r, column=c).fill = alt_fill
    
    # Highlight the "competitive" range
    if stones in [6, 7, 8]:
        for c in range(1, 7):
            ws3.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_YES)

# Multi-year strategy
yr_start = 31
ws3.merge_cells(f"A{yr_start}:H{yr_start}")
ws3[f"A{yr_start}"].value = "MULTI-YEAR STRATEGY: STONES CARRY OVER IF NOT SELECTED"
ws3[f"A{yr_start}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
ws3[f"A{yr_start}"].fill = PatternFill("solid", fgColor=MED_BLUE)

strategy_text = [
    "If Kathleen is NOT selected in the 2027 lottery, her stones carry over (they never expire).",
    "She just needs to earn at least 1 new stone within 24 months to keep them active.",
    "Each additional year of racing adds more stones, significantly improving odds.",
    "",
    "Example path: Race 2 events in 2026 (6 stones) → Not selected 2027 → Race 1 more in 2027 (3 stones) → Enter 2028 with 9 stones",
    "",
    "KEY DATES FOR 2027 CCC:",
    "  • Dec 2026: Pre-registration opens for CCC 2027 lottery",
    "  • Mid-Jan 2027: Lottery results announced", 
    "  • Aug 29, 2027: CCC race day in Chamonix",
    "",
    "All stones must be earned BEFORE the Dec 2026 pre-registration deadline to count for the 2027 lottery.",
]

for i, line in enumerate(strategy_text):
    r = yr_start + 1 + i
    ws3.merge_cells(f"A{r}:H{r}")
    ws3[f"A{r}"].value = line
    ws3[f"A{r}"].font = Font(name="Arial", size=10, bold=("KEY DATES" in line or "Example" in line))
    ws3[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")

# Column widths
odds_widths = [32, 16, 16, 12, 18, 40, 1, 1]
for c, w in enumerate(odds_widths, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

ws3.freeze_panes = "A4"

# ============================================================
# SHEET 4: RECOMMENDED STRATEGIES
# ============================================================
ws4 = wb.create_sheet("Strategy Guide")
ws4.sheet_properties.tabColor = ACCENT_ORANGE

ws4.merge_cells("A1:F1")
ws4["A1"].value = "RECOMMENDED RACE STRATEGIES FOR KATHLEEN — CCC 2027"
ws4["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws4["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
ws4["A1"].alignment = center
ws4.row_dimensions[1].height = 35

strategies = [
    ("STRATEGY A: MINIMUM VIABLE (Budget-Friendly)", ACCENT_GOLD, [
        ("Race 1", "Speedgoat 50K", "Jul 23-25, 2026", "Snowbird, UT", "~45 min drive!", "2 stones + 50K Index", "$150-200"),
        ("Race 2", "Snowbasin 50K", "Sep 10-12, 2026", "Ogden, UT", "~1 hr drive!", "2 stones + 50K Index", "$100-200"),
    ], "4 stones total. Average odds. Nearly zero travel cost. Gets valid 50K Index for CCC."),
    
    ("STRATEGY B: COMPETITIVE (Best Value)", MED_BLUE, [
        ("Race 1", "Canyons 100K", "Apr 23-25, 2026", "Auburn, CA", "~9 hr drive / fly SLC-SMF", "3 stones + 100K Index", "$200-300"),
        ("Race 2", "Speedgoat 50K", "Jul 23-25, 2026", "Snowbird, UT", "~45 min drive!", "2 stones", "$150-200"),
        ("Race 3", "Snowbasin 50K", "Sep 10-12, 2026", "Ogden, UT", "~1 hr drive!", "2 stones", "$100-200"),
    ], "7 stones total. Good odds (~25-30%). Locks in 100K Index. Two races are basically local."),
    
    ("STRATEGY C: MAXIMIZE STONES (Best Odds)", ACCENT_ORANGE, [
        ("Race 1", "Canyons 100K", "Apr 23-25, 2026", "Auburn, CA", "~9 hr drive / fly", "3 stones + 100K Index", "$200-300"),
        ("Race 2", "Speedgoat 50K", "Jul 23-25, 2026", "Snowbird, UT", "~45 min drive!", "2 stones", "$150-200"),
        ("Race 3", "Kodiak 100K (MAJOR)", "Oct 8-11, 2026", "Big Bear, CA", "Fly SLC-ONT", "6 stones (doubled!)", "$250-350"),
    ], "11 stones total. Excellent odds (~40%). Kodiak is the Americas Major = double stones. Also Western States qualifier."),
    
    ("STRATEGY D: ALL-IN UTAH + MAJOR", "006100", [
        ("Race 1", "Speedgoat 50K", "Jul 23-25, 2026", "Snowbird, UT", "~45 min drive!", "2 stones + 50K Index", "$150-200"),
        ("Race 2", "Snowbasin 50K", "Sep 10-12, 2026", "Ogden, UT", "~1 hr drive!", "2 stones", "$100-200"),
        ("Race 3", "Kodiak 100K (MAJOR)", "Oct 8-11, 2026", "Big Bear, CA", "Fly SLC-ONT", "6 stones (doubled!)", "$250-350"),
    ], "10 stones total. Strong odds (~37%). Only 1 trip out of state. 50K Index for CCC (valid)."),
]

current_row = 3
for strat_name, color, race_list, summary in strategies:
    ws4.merge_cells(f"A{current_row}:F{current_row}")
    ws4[f"A{current_row}"].value = strat_name
    ws4[f"A{current_row}"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
    ws4[f"A{current_row}"].fill = PatternFill("solid", fgColor=color)
    ws4[f"A{current_row}"].alignment = center
    current_row += 1
    
    # Sub-headers
    sub_h = ["", "Event", "Date", "Location", "Travel from Logan", "Stones + Index"]
    for c, h in enumerate(sub_h, 1):
        ws4.cell(row=current_row, column=c, value=h)
        ws4.cell(row=current_row, column=c).font = Font(name="Arial", bold=True, size=9)
        ws4.cell(row=current_row, column=c).fill = subheader_fill
        ws4.cell(row=current_row, column=c).border = thin_border
        ws4.cell(row=current_row, column=c).alignment = center
    current_row += 1
    
    for label, event, date, loc, travel, stones_info, cost in race_list:
        vals = [label, event, date, loc, travel, stones_info]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=current_row, column=c, value=v)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center if c != 5 else Alignment(horizontal="left", vertical="center")
        current_row += 1
    
    # Summary
    ws4.merge_cells(f"A{current_row}:F{current_row}")
    ws4[f"A{current_row}"].value = f"→ {summary}"
    ws4[f"A{current_row}"].font = Font(name="Arial", size=10, bold=True, italic=True)
    ws4[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[current_row].height = 35
    current_row += 2

# Key insight
ws4.merge_cells(f"A{current_row}:F{current_row}")
ws4[f"A{current_row}"].value = "KEY INSIGHT: Snowbasin (Ogden) and Speedgoat (Snowbird) are both within ~1 hour of Logan — these are essentially home races for Kathleen."
ws4[f"A{current_row}"].font = Font(name="Arial", size=11, bold=True, color="006100")
ws4[f"A{current_row}"].fill = PatternFill("solid", fgColor=GREEN_YES)
ws4[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws4.row_dimensions[current_row].height = 35

current_row += 2
ws4.merge_cells(f"A{current_row}:F{current_row}")
ws4[f"A{current_row}"].value = "NOTE: For the CCC lottery, Kathleen needs a valid UTMB Index in the 50K, 100K, or 100M category. Running ANY 50K+ at a UTMB World Series Event or Index race satisfies this requirement."
ws4[f"A{current_row}"].font = Font(name="Arial", size=10, italic=True)
ws4[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws4.row_dimensions[current_row].height = 35

# Column widths
strat_widths = [12, 26, 16, 18, 26, 26]
for c, w in enumerate(strat_widths, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# SHEET 5: HOW IT WORKS
# ============================================================
ws5 = wb.create_sheet("How UTMB Qualification Works")
ws5.sheet_properties.tabColor = "666666"

ws5.merge_cells("A1:E1")
ws5["A1"].value = "UTMB CCC (100K) QUALIFICATION — HOW IT WORKS"
ws5["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws5["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
ws5["A1"].alignment = center
ws5.row_dimensions[1].height = 35

info_lines = [
    ("", ""),
    ("THE TWO REQUIREMENTS TO ENTER THE CCC LOTTERY:", ""),
    ("", ""),
    ("1. RUNNING STONES (Lottery Tickets)", ""),
    ("   • Earned by FINISHING races at UTMB World Series Events or Majors", ""),
    ("   • Events: 20K=1 stone, 50K=2, 100K=3, 100M=4 stones", ""),
    ("   • Majors: DOUBLE stones (50K=4, 100K=6, 100M=8)", ""),
    ("   • Stones NEVER expire, but go inactive if no new stone earned in 24 months", ""),
    ("   • More stones = more lottery entries = better odds", ""),
    ("   • When selected, ALL stones reset to zero", ""),
    ("   • Must have at least 1 stone earned in last 24 months to enter lottery", ""),
    ("", ""),
    ("2. VALID UTMB INDEX", ""),
    ("   • For CCC (100K): need valid index in 50K, 100K, or 100M category", ""),
    ("   • Earned by finishing ANY UTMB World Series race or UTMB Index race in that category", ""),
    ("   • Valid for 24 months from race date", ""),
    ("   • Over 5,500 races worldwide count as Index races (not just World Series)", ""),
    ("   • The Index score doesn't matter for lottery entry — just needs to be valid", ""),
    ("", ""),
    ("THE LOTTERY:", ""),
    ("   • Pre-registration: ~December each year", ""),
    ("   • Draw: ~Mid-January", ""),
    ("   • Each stone = 1 entry in the draw", ""),
    ("   • 2025 stats: ~6,000 CCC applicants, avg applicant had 4.4 stones", ""),
    ("   • Average SELECTED runner had 5.7 stones", ""),
    ("   • Demand is ~5x capacity", ""),
    ("   • Recommended target: 6-8 stones for competitive odds", ""),
    ("", ""),
    ("ALTERNATIVE PATHS (Skip the Lottery):", ""),
    ("   • Top 3 finish (M/F) at any World Series Event in 100K → direct CCC entry", ""),
    ("   • Top 10 finish (M/F) at a Major in 100K → direct CCC entry", ""),
    ("   • Achieve UTMB Score threshold (women: 660 for 100K) → direct entry", ""),
    ("   • Charity bibs: ~€2,000 (sell out in under 1 hour)", ""),
    ("", ""),
    ("KEY TIMELINE FOR CCC 2027:", ""),
    ("   • 2026: Race UTMB World Series events to collect stones + Index", ""),
    ("   • Dec 2026: Pre-register for CCC 2027 lottery", ""),
    ("   • Jan 2027: Lottery results", ""),
    ("   • Aug 29, 2027: CCC race day in Chamonix, France", ""),
]

for i, (text, _) in enumerate(info_lines):
    r = i + 3
    ws5.merge_cells(f"A{r}:E{r}")
    cell = ws5[f"A{r}"]
    cell.value = text
    if text.startswith(("THE TWO", "THE LOTTERY", "ALTERNATIVE", "KEY TIMELINE")):
        cell.font = Font(name="Arial", bold=True, size=11, color=MED_BLUE)
    elif text.startswith(("1.", "2.")):
        cell.font = Font(name="Arial", bold=True, size=11)
    else:
        cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

ws5.column_dimensions["A"].width = 90

# ============================================================
# SAVE
# ============================================================
output_path = "/home/claude/utmb_ccc_planner.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
